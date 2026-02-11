"""Generate an XLSX workbook from consolidated rows."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

COLUMNS = [
    ("Tipo", 8),
    ("Periodo", 16),
    ("Conta", 45),
    ("Mascara_Contabil", 20),
    ("Conta_Padronizada", 45),
    ("Sinal", 6),
    ("Classificacao_Padrao", 40),
    ("Ano_Anterior", 18),
    ("Ano_Atual", 18),
    ("Pagina_Origem", 16),
]


def build_xlsx(rows: list[dict], errors: list[dict] | None = None) -> bytes:
    """Create an XLSX file in memory and return its bytes.

    Args:
        rows: Consolidated row dicts with keys matching COLUMNS.
        errors: Optional list of extraction errors (will be used in a future update).

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Balancete Consolidado"

    # --- Header row ---
    for col_idx, (col_name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    # --- Data rows ---
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Number formatting for monetary columns
            if col_name in ("Ano_Anterior", "Ano_Atual"):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Auto-filter
    num_cols = len(COLUMNS)
    last_col_letter = chr(ord("A") + num_cols - 1)
    if rows:
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Error report sheet ---
    ws_errors = wb.create_sheet("Relatório de Erros")

    # Headers
    error_columns = [("Página", 20), ("Erro", 100)]
    for col_idx, (col_name, width) in enumerate(error_columns, start=1):
        cell = ws_errors.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws_errors.column_dimensions[cell.column_letter].width = width

    # Data
    if errors:
        for row_idx, error in enumerate(errors, start=2):
            ws_errors.cell(row=row_idx, column=1, value=error.get("pagina", ""))
            ws_errors.cell(row=row_idx, column=2, value=error.get("erro", ""))
    else:
        ws_errors.cell(row=2, column=1, value="")
        ws_errors.cell(row=2, column=2, value="Nenhum erro encontrado")

    ws_errors.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
